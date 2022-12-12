import sys, os
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage: " + sys.argv[0] + " <Skill and Knowledge Assessment.xlsx>")
    sys.exit(1)

if not os.path.exists(sys.argv[1]):
    print("File not found: " + sys.argv[1])
    sys.exit(2)

wb = load_workbook(filename=sys.argv[1])

ws = wb['HW']
i = 2
print('INSERT INTO hardware (product) VALUES')
while True:
    c = ws['A' + str(i)]
    if not c.value:
        print()
        break
    elif i > 2:
        print(',')
    print('(\'' + c.value + '\')', end='')
    i += 1
print('ON CONFLICT DO NOTHING;')

i = "A"
print('INSERT INTO tasks_hw (task) VALUES')
while True:
    c = ws[i + '1']
    if not c.value:
        print()
        break
    elif ord(i) > ord('A'):
        print(', ')
    print('(\'' + c.value + '\')', end='')
    i = chr(ord(i) + 1)
print('ON CONFLICT DO NOTHING;')


ws = wb['SW']
i = 2
print('INSERT INTO software (product) VALUES')
while True:
    c = ws['A' + str(i)]
    if not c.value:
        print()
        break
    elif i > 2:
        print(',')
    print('(\'' + c.value + '\')', end='')
    i += 1
print('ON CONFLICT DO NOTHING;')

i = "A"
print('INSERT INTO tasks_sw (task) VALUES')
while True:
    c = ws[i + '1']
    if not c.value:
        print()
        break
    elif ord(i) > ord('A'):
        print(', ')
    print('(\'' + c.value + '\')', end='')
    i = chr(ord(i) + 1)
print('ON CONFLICT DO NOTHING;')

ws = wb['Process']
i = 2
print('INSERT INTO processes (process) VALUES')
while True:
    c = ws['A' + str(i)]
    if not c.value:
        print()
        break
    elif i > 2:
        print(',')
    print('(\'' + c.value + '\')', end='')
    i += 1
print('ON CONFLICT DO NOTHING;')
