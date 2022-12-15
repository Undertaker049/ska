import sys, os
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage: python3" + sys.argv[0] + " <Employee-Skills and Knowledge Assessment-Year-Month.xlsx>")
    sys.exit(1)

if not os.path.exists(sys.argv[1]):
    print("File not found: " + sys.argv[1])
    sys.exit(2)

(name, x, year, month) = sys.argv[1].replace('.xlsx', '').split('-')

#employee = input(f'Enter employee name [{name}]: ')
#if not employee:
employee = name
print(f'INSERT INTO employees (name) VALUES (\'{employee}\') ON CONFLICT DO NOTHING;')

wb = load_workbook(filename=sys.argv[1])

ws = wb['HW']
i = 2
while (skill := ws['A' + str(i)]) and skill.value:
    j = "B"
    while (task := ws[j + '1']) and task.value:
        c = ws[j + str(i)]
        if not c.value:
            level = 'None'
        else:
            level = c.value.split(' — ')[1]
        if not level == 'None':
            print(f'INSERT INTO skills_hw (timestamp, employee, product, task, level) SELECT date(), e.id, hw.id, t.id, l.id FROM employees e INNER JOIN hardware hw ON hw.product = \'{skill.value}\' INNER JOIN tasks_hw t ON t.task = \'{task.value}\' INNER JOIN levels l ON l.level = \'{level}\' WHERE e.name = \'{employee}\';')
        j = chr(ord(j) + 1)
    i += 1

ws = wb['SW']
i = 2
while (skill := ws['A' + str(i)]) and skill.value:
    j = "B"
    while (task := ws[j + '1']) and task.value:
        c = ws[j + str(i)]
        if not c.value:
            level = 'None'
        else:
            level = c.value.split(' — ')[1]
        if not level == 'None':
            print(f'INSERT INTO skills_sw (timestamp, employee, product, task, level) SELECT date(), e.id, sw.id, t.id, l.id FROM employees e INNER JOIN software sw ON sw.product = \'{skill.value}\' INNER JOIN tasks_sw t ON t.task = \'{task.value}\' INNER JOIN levels l ON l.level = \'{level}\' WHERE e.name = \'{employee}\';')
        j = chr(ord(j) + 1)
    i += 1

ws = wb['Process']
i = 2
while (skill := ws['A' + str(i)]) and skill.value:
    c = ws['B' + str(i)]
    if not c.value:
        level = 'None'
    else:
        level = c.value.split(' — ')[1]
    if not level == 'None':
        print(f'INSERT INTO skills_pr (timestamp, employee, process, level) SELECT date(), e.id, pr.id, l.id FROM employees e INNER JOIN processes pr ON pr.process = \'{skill.value}\' INNER JOIN levels l ON l.level = \'{level}\' WHERE e.name = \'{employee}\';')
    i += 1
